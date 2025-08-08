import os
from datetime import datetime

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter


RUNS_HEADERS = [
    "RunID",
    "RunDate",
    "Company",
    "DateRange",
    "TotalRows",
    "SuccessClassic",
    "FailedClassic",
    "SuccessRelaxed",
    "FailedRelaxed",
    "ReportPath",
]

PG_SUMMARY_HEADERS = [
    "RunID",
    "Company",
    "PG_NAME",
    "TotalRows",
    "Order_TRUE",
    "Order_FALSE",
    "Order_SKIPPED",
    "Patient_TRUE",
    "Patient_FALSE",
    "Patient_SKIPPED",
    "PDF_TRUE",
    "PDF_FALSE",
    "PDF_SKIPPED",
    "SuccessClassic",
    "SuccessRelaxed",
    "TopFailureReason",
]


def style_header(row_cells):
    header_fill = PatternFill("solid", fgColor="F0F3F5")
    bold_font = Font(bold=True)
    thin = Side(border_style="thin", color="D0D7DE")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    for cell in row_cells:
        cell.font = bold_font
        cell.alignment = Alignment(vertical="center")
        cell.fill = header_fill
        cell.border = border


def autosize(ws, min_width=10, max_width=42):
    for col in ws.columns:
        col_letter = get_column_letter(col[0].column)
        max_len = 0
        for cell in col:
            value = cell.value
            if value is None:
                continue
            try:
                cell_len = len(str(value))
            except Exception:
                cell_len = 0
            max_len = max(max_len, cell_len)
        ws.column_dimensions[col_letter].width = min(max(max_len + 2, min_width), max_width)


def build_runs_sheet(wb):
    ws = wb.create_sheet("Runs")
    ws.append(RUNS_HEADERS)
    style_header(ws[1])
    ws.freeze_panes = "A2"
    autosize(ws)
    return ws


def build_pg_summary_sheet(wb):
    ws = wb.create_sheet("PG_Summary")
    ws.append(PG_SUMMARY_HEADERS)
    style_header(ws[1])
    ws.freeze_panes = "A2"
    autosize(ws)
    return ws


def build_dashboard_sheet(wb):
    ws = wb.create_sheet("Dashboard")

    title_font = Font(bold=True, size=14)
    section_font = Font(bold=True)

    ws["A1"] = "Doctor Alliance – PG Tracking Dashboard"
    ws["A1"].font = title_font

    # Filters
    ws["A3"] = "Filters"
    ws["A3"].font = section_font
    ws["A4"] = "RunID"
    ws["B4"] = ""  # user input
    ws["A5"] = "Company"
    ws["B5"] = ""  # user input

    # KPIs
    ws["A7"] = "KPIs"
    ws["A7"].font = section_font
    ws["A8"] = "Total Rows"
    ws["B8"] = "=SUMIFS(PG_Summary!D:D,PG_Summary!A:A,B4,PG_Summary!B:B,B5)"

    ws["A9"] = "SuccessClassic"
    ws["B9"] = "=SUMIFS(PG_Summary!N:N,PG_Summary!A:A,B4,PG_Summary!B:B,B5)"
    ws["A10"] = "SuccessClassic %"
    ws["B10"] = "=IFERROR(B9/B8,0)"

    ws["A11"] = "SuccessRelaxed"
    ws["B11"] = "=SUMIFS(PG_Summary!O:O,PG_Summary!A:A,B4,PG_Summary!B:B,B5)"
    ws["A12"] = "SuccessRelaxed %"
    ws["B12"] = "=IFERROR(B11/B8,0)"

    # Instructions for PG breakdown and charting
    ws["A14"] = "How to use"
    ws["A14"].font = section_font
    ws["A15"] = (
        "1) Set RunID and Company above.\n"
        "2) After each run, paste per-PG rows into 'PG_Summary' (one row per PG for the RunID).\n"
        "   Columns expected: TotalRows, Order/PATIENT/PDF status counts, SuccessClassic, SuccessRelaxed.\n"
        "3) The KPIs update automatically.\n"
        "4) Create a PivotChart on 'PG_Summary' (Rows=PG_NAME, Values=SuccessRelaxed & Failures) and place it here."
    )
    ws.merge_cells("A15:F18")
    ws["A15"].alignment = Alignment(wrap_text=True, vertical="top")

    # Reserve area for PG breakdown table pasted from a Pivot
    ws["A20"] = "PG Breakdown (paste a Pivot here)"
    ws["A20"].font = section_font
    ws["A21"] = "PG_NAME"
    ws["B21"] = "SuccessRelaxed"
    ws["C21"] = "Failures (TotalRows - SuccessRelaxed)"
    style_header(ws[21])

    autosize(ws)
    return ws


def create_tracker(path: str = "PG_Tracker.xlsx", overwrite: bool = True) -> str:
    if os.path.exists(path) and not overwrite:
        return path

    wb = Workbook()
    # Remove default sheet
    default = wb.active
    wb.remove(default)

    build_runs_sheet(wb)
    build_pg_summary_sheet(wb)
    build_dashboard_sheet(wb)

    # Metadata sheet (optional, hidden)
    meta = wb.create_sheet("_Meta")
    meta["A1"] = "Generated"
    meta["B1"] = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    meta.sheet_state = "hidden"

    wb.save(path)
    return path


if __name__ == "__main__":
    out = create_tracker()
    print(f"Created: {out}")


