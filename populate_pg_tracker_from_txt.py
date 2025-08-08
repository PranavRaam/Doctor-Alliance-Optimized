import os
import re
from typing import List, Dict, Optional

from openpyxl import load_workbook
from openpyxl.workbook.workbook import Workbook


UUID_RE = re.compile(r"[0-9a-fA-F]{8}-[0-9a-fA-F]{4}-[0-9a-fA-F]{4}-[0-9a-fA-F]{4}-[0-9a-fA-F]{12}")


def is_region_header(line: str) -> bool:
    s = line.strip()
    if not s:
        return False
    if set(s) <= set("-_"):
        return False
    # Likely region headers: all uppercase letters/spaces and short
    return s == s.upper() and len(s) <= 25 and " " in s or s in {"WEST", "CENTRAL", "EAST", "EAST CENTRAL"}


def extract_uuid(text: str) -> Optional[str]:
    m = UUID_RE.search(text)
    return m.group(0) if m else None


def extract_helper_id(text_after_uuid: str) -> str:
    # Heuristics: pick the last alphanumeric token (often like ihelperph..., dallianceph..., jmonroyph, etc.)
    # Remove common separators
    tail = text_after_uuid.strip().strip("-:")
    if not tail:
        return ""
    # Try to find tokens; prefer ones ending/containing 'ph'
    tokens = re.findall(r"[A-Za-z][A-Za-z0-9]*", tail)
    if not tokens:
        return ""
    # Prefer token containing 'ph'
    for tok in reversed(tokens):
        if 'ph' in tok.lower():
            return tok
    return tokens[-1]


def parse_pg_list(path: str) -> List[Dict[str, str]]:
    rows: List[Dict[str, str]] = []
    current_region = ""
    seen_keys = set()

    with open(path, "r", encoding="utf-8") as f:
        for raw in f:
            line = raw.strip()
            if not line:
                continue
            if set(line) <= set("-_"):
                continue
            if is_region_header(line):
                current_region = line.strip()
                continue

            # Try to parse a PG row
            uuid = extract_uuid(line)
            name_part = line
            helper_id = ""
            if uuid:
                # Name precedes the UUID, helper likely after
                before, _, after = line.partition(uuid)
                # Remove trailing ' -' and spaces in name part
                name_part = before.strip().rstrip("-:").strip()
                helper_id = extract_helper_id(after)
            else:
                # No UUID: try to split on last '-' and get helper
                if '-' in line:
                    # Use rpartition to split from right
                    before, sep, after = line.rpartition('-')
                    name_part = before.strip()
                    helper_id = extract_helper_id(after)
                else:
                    name_part = line.strip()
                    helper_id = ""

            # Normalize name (remove trailing/leading punctuation)
            name = name_part.strip().strip('-').strip()
            if not name:
                continue

            key = (uuid or name).lower()
            if key in seen_keys:
                continue
            seen_keys.add(key)

            rows.append({
                "region": current_region,
                "pg_name": name,
                "pg_company_id": uuid or "",
                "helper_id": helper_id or "",
            })

    return rows


def ensure_tracker(filepath: str) -> Workbook:
    if not os.path.exists(filepath):
        # Lazy import to avoid circulars if not needed
        from create_pg_tracker import create_tracker
        create_tracker(filepath, overwrite=True)
    wb = load_workbook(filepath)
    return wb


PG_MASTER_HEADERS = ["Region", "PG_Name", "PG_Company_ID", "Helper_ID"]


def write_pg_master(wb: Workbook, items: List[Dict[str, str]]):
    if "PG_Master" in wb.sheetnames:
        ws = wb["PG_Master"]
        wb.remove(ws)
    ws = wb.create_sheet("PG_Master", 0)
    ws.append(PG_MASTER_HEADERS)
    for it in items:
        ws.append([it["region"], it["pg_name"], it["pg_company_id"], it["helper_id"]])


def write_pg_summary_template(wb: Workbook, items: List[Dict[str, str]]):
    # Prefill PG_Summary with a template row per PG (blank RunID/Company, zero counts)
    headers = [
        "RunID",
        "Company",
        "PG_NAME",
        "TotalRows",
        "Order_TRUE",
        "Order_FALSE",
        "Order_SKIPPED",
        "Patient_TRUE",
        "Patient_FALSE",
        "Patient_SKIPPED",
        "PDF_TRUE",
        "PDF_FALSE",
        "PDF_SKIPPED",
        "SuccessClassic",
        "SuccessRelaxed",
        "TopFailureReason",
    ]
    if "PG_Summary" in wb.sheetnames:
        ws = wb["PG_Summary"]
        wb.remove(ws)
    ws = wb.create_sheet("PG_Summary", 1)
    ws.append(headers)
    for it in items:
        ws.append([
            "",  # RunID
            "",  # Company
            it["pg_name"],
            0,  # TotalRows
            0, 0, 0,  # Order_* counts
            0, 0, 0,  # Patient_* counts
            0, 0, 0,  # PDF_* counts
            0,        # SuccessClassic
            0,        # SuccessRelaxed
            "",      # TopFailureReason
        ])


def main():
    src = "PG_IDS_HELPER_IDS.txt"
    dst = "PG_Tracker.xlsx"
    if not os.path.exists(src):
        raise FileNotFoundError(f"Not found: {src}")
    items = parse_pg_list(src)
    wb = ensure_tracker(dst)
    write_pg_master(wb, items)
    write_pg_summary_template(wb, items)
    wb.save(dst)
    print(f"Updated tracker with {len(items)} PG rows: {dst}")


if __name__ == "__main__":
    main()


